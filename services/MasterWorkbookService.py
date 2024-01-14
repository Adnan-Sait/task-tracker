from typing import Dict
import pandas as pd
import os

from pandas.io.excel._base import ExcelFile

from services.DataRepository import DataRepository
from models.DailyActivity import DailyActivity
from models.DayOverview import DayOverview
from table_models.ProjectTable import ProjectTable
from table_models.TaskTable import TaskTable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class MasterWorkbookService(DataRepository):

    def __init__(self, filePath: str):
        self.filePath = os.path.abspath(filePath)
        self.excelFile = pd.ExcelFile(filePath)
        self.projectNameProjectListDict = self.__getProjectNameProjectListDictionary(
            self.excelFile, "Project")
        self.taskIdTaskDetailsDict = self.__getTaskIdTaskDetailsDictionary(
            self.excelFile, "Task")
        self.projectIdTaskListDict = self.__getProjectIdTaskListDictionary(
            self.excelFile, "Task")
        self.morningScheduleIdDetailsDict = self.__getMorningScheduleIdDetailsDictionary(
            self.excelFile, "MorningSchedule")
        self.taskTrackerMasterFile = load_workbook(filePath)

        self.currentDailyActivityRow = self.__getTotalRowsCount(
            "DailyActivity", "dateRecorded") + 1
        self.currentDayOverviewRow = self.__getTotalRowsCount(
            "DayOverview", "dateRecorded") + 1

    def __getTotalRowsCount(self, sheetName: str, headerName: str) -> int:
        sheetData = self.excelFile.parse(sheetName, 0)
        return len(sheetData[headerName])

    def __del__(self):
        self.taskTrackerMasterFile.close()
        self.excelFile.close()

    def saveDailyActivity(self, dailyActivity: DailyActivity) -> DailyActivity:
        dailyActivitySheet = self.taskTrackerMasterFile["DailyActivity"]

        self.currentDailyActivityRow += 1
        column = 1

        # write to master sheet
        dailyActivitySheet.cell(
            row=self.currentDailyActivityRow, column=column).value = dailyActivity.getDate()
        column += 1
        dailyActivitySheet.cell(row=self.currentDailyActivityRow, column=column).value = self.taskIdTaskDetailsDict.get(
            dailyActivity.getTaskId()).taskName
        column += 1
        dailyActivitySheet.cell(
            row=self.currentDailyActivityRow, column=column).value = dailyActivity.getEffort()
        column += 1
        dailyActivitySheet.cell(row=self.currentDailyActivityRow,
                                column=column).value = dailyActivity.getDescription()

        return dailyActivity

    def saveBulkDailyActivity(self, dailyActivityList: list[DailyActivity]) -> list[DailyActivity]:
        dailyActivitySheet = self.taskTrackerMasterFile["DailyActivity"]

        # write to master sheet
        for dailyActivity in dailyActivityList:
            self.currentDailyActivityRow += 1
            column = 1
            dailyActivitySheet.cell(
                row=self.currentDailyActivityRow, column=column).value = dailyActivity.getDate()
            column += 1
            dailyActivitySheet.cell(row=self.currentDailyActivityRow, column=column).value = self.taskIdTaskDetailsDict.get(
                dailyActivity.getTaskId()).taskName
            column += 1
            dailyActivitySheet.cell(
                row=self.currentDailyActivityRow, column=column).value = dailyActivity.getEffort()
            column += 1
            dailyActivitySheet.cell(row=self.currentDailyActivityRow,
                                    column=column).value = dailyActivity.getDescription()

        return dailyActivityList

    def getTaskId(self, projectId: int, taskName: str) -> int:
        taskList = self.projectIdTaskListDict.get(projectId)
        if (taskList):
            for task in taskList:
                if (task.taskName.lower() == taskName.lower()):
                    return task.taskId
        raise Exception(
            f"No task with projectId, {projectId} and taskName, {taskName}")

    def getProjectId(self, projectName: str) -> int:
        project = self.projectNameProjectListDict.get(projectName.lower())
        if (project):
            return project.projectId
        else:
            raise Exception(f"No project with projectName, {projectName}")

    def saveDayOverview(self, dayOverview: DayOverview) -> bool:
        dayOverviewSheet = self.taskTrackerMasterFile["DayOverview"]

        self.currentDayOverviewRow += 1
        column = 1

        # write to master sheet
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.dateRecorded
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.startTime
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.leavingTime
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.workHours
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.breakHours
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.lunchDuration
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = dayOverview.prayer
        column += 1
        dayOverviewSheet.cell(row=self.currentDayOverviewRow,
                              column=column).value = self.morningScheduleIdDetailsDict[dayOverview.morningScheduleId]["morningScheduleRepresentation"]

        return dayOverview

    # TODO
    def getProjectInfo(self):
        pass

    def getActiveProjects(self) -> list:
        activeProjects = []
        for project in self.projectNameProjectListDict.items():
            if (project[1].status == "ACTIVE"):
                activeProjects.append(project[1])
        if (len(activeProjects) == 0):
            return None
        activeProjects.sort(key=lambda x: x.projectName)
        return activeProjects

    def getActiveTasksByProjectId(self, projectId: int) -> list:
        activeTasks = []
        taskList = self.projectIdTaskListDict.get(projectId)
        if (taskList is None):
            raise Exception(f"No tasks linked with projectId, {projectId}")
        for task in taskList:
            if (task.status == "ACTIVE"):
                activeTasks.append(task)
        activeTasks.sort(key=lambda x: x.taskName)
        return activeTasks

    def getTaskTypes(self) -> list:
        tasksTypeSheetData = self.excelFile.parse("TaskType", 0)
        taskTypeList = []
        for index in range(len(tasksTypeSheetData["taskTypeName"])):
            taskTypeList.append(tasksTypeSheetData["taskTypeName"][index])
        return taskTypeList

    def commit(self):
        self.taskTrackerMasterFile.save(self.filePath)

    def __getProjectNameProjectListDictionary(self, excelFile: ExcelFile, projectSheetName: str) -> dict:
        projectSheetData = excelFile.parse(projectSheetName, 0)
        projectIdProjectDict = {}
        for index in range(len(projectSheetData["projectName"])):
            project = ProjectTable()
            project.projectId = projectSheetData["projectId"][index]
            project.projectName = projectSheetData["projectName"][index]
            project.startTimestamp = projectSheetData["startTimestamp"][index]
            project.status = projectSheetData["status"][index]
            project.description = projectSheetData["description"][index]
            projectIdProjectDict[project.projectName.lower()] = project
        return projectIdProjectDict

    def __getProjectIdTaskListDictionary(self, excelFile: ExcelFile, taskSheetName: str) -> dict:
        taskSheetData = excelFile.parse(taskSheetName, 0)
        projectIdTaskListDict = {}
        taskList = []
        for index in range(len(taskSheetData["taskName"])):
            task = TaskTable()
            task.taskId = taskSheetData["taskId"][index]
            task.taskName = taskSheetData["taskName"][index]
            task.projectName = taskSheetData["projectName"][index]
            task.status = taskSheetData["status"][index]
            task.startTimestamp = taskSheetData["startTimestamp"][index]
            task.completeTimestamp = taskSheetData["completeTimestamp"][index]
            task.description = taskSheetData["description"][index]
            projectId = self.getProjectId(task.projectName)
            # projectIdTaskListDict[projectId] = task

            if (projectIdTaskListDict.get(projectId)):
                taskList = projectIdTaskListDict[projectId]
                taskList.append(task)
            else:
                taskList = []
                taskList.append(task)
            projectIdTaskListDict[projectId] = taskList
        return projectIdTaskListDict

    def __getTaskIdTaskDetailsDictionary(self, excelFile: ExcelFile, taskSheetName: str) -> dict:
        taskSheetData = excelFile.parse(taskSheetName, 0)
        taskIdTaskDetailsDict = {}
        for index in range(len(taskSheetData["taskId"])):
            task = TaskTable()
            task.taskId = taskSheetData["taskId"][index]
            task.taskName = taskSheetData["taskName"][index]
            task.projectName = taskSheetData["projectName"][index]
            task.status = taskSheetData["status"][index]
            task.startTimestamp = taskSheetData["startTimestamp"][index]
            task.completeTimestamp = taskSheetData["completeTimestamp"][index]
            task.description = taskSheetData["description"][index]
            taskIdTaskDetailsDict[task.taskId] = task
        return taskIdTaskDetailsDict

    def __getMorningScheduleIdDetailsDictionary(self, excelFile: ExcelFile, morningScheduleSheetName: str) -> dict:
        morningScheduleIdDetailsDict = {}
        morningScheduleSheetData = excelFile.parse(morningScheduleSheetName, 0)

        for index in range(len(morningScheduleSheetData["morningScheduleId"])):
            morningScheduleObj = {}
            morningScheduleObj["morningScheduleId"] = morningScheduleSheetData["morningScheduleId"][index]
            morningScheduleObj["morningScheduleRepresentation"] = morningScheduleSheetData["morningScheduleRepresentation"][index]
            morningScheduleObj["description"] = morningScheduleSheetData["description"][index]
            morningScheduleIdDetailsDict[morningScheduleObj["morningScheduleId"]
                                         ] = morningScheduleObj
        return morningScheduleIdDetailsDict
