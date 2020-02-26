import pandas as pd
import yaml
import datetime
import time
import math
import os
import re
import shutil
from models.ExcelDataConsolidation import ExcelDataConsolidation
from models.DailyActivity import DailyActivity
from models.DayOverview import DayOverview
from models.Project import Project
from models.Task import Task
import database.ExcelDatabase as excelDatabase
from utilities.Constants import Constants
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from pandas.io.excel._base import ExcelFile,DataFrame

from openpyxl import load_workbook

with open('./resources/config.yml') as configFile:
    configurationValues = yaml.load(configFile, Loader=yaml.BaseLoader)

FILE_PATH = configurationValues['filePath']
FILE_DESTINATION_PATH = configurationValues['fileDestinationPath']
TASK_TRACKER_TEMPLATE_PATH = configurationValues['taskTrackerTemplatePath']
DATE_RECORDED_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['date']
CATEGORY_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['category']
PROJECT_NAME_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['project']
TASK_NAME_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['task']
TYPE_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['type']
START_TIME_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['startTime']
END_TIME_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['endTime']
DURATION_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['duration']
DESCRIPTION_SHEET_IDENTIFIER = configurationValues['excelConfiguration']['dataTableColumns']['description']
CATEGORY_CONST_DICT = configurationValues['excelConfiguration']['categoryDict']
PRAYER_DEFAULT = configurationValues['defaultPrayer']
CELL_IDENTIFIER = configurationValues['excelConfiguration']['cellAddress']
DEFAULT_PROJECT = configurationValues['defaultProject']


def parseUnixTime(fileName: str) -> float:
    """Parses the Unix timestamp from the file name
    
    Arguments:
        fileName {str} -- [description]
    
    Returns:
        [float] -- [description]
    """
    currentDate = getDateRecordedFromFileName(fileName)
    unixTime = time.mktime(datetime.datetime.strptime(currentDate.strftime("%d/%m/%Y"), "%d/%m/%Y").timetuple())
    return unixTime

def getFile(filename: str) -> ExcelFile:
    """Reads Data from an Excel Sheet
    
    Arguments:
        filename {str} -- [description]
    
    Returns:
        ExcelFile -- [description]
    """
    excelFile = pd.ExcelFile(FILE_PATH+'/'+filename)
    return excelFile

def deleteFile(filename: str):
    """Deletes the specified filename from the preconfigured directory
    
    Arguments:
        filename {str} -- [description]
    """
    shutil.move(FILE_PATH+'/'+filename, FILE_DESTINATION_PATH+"/"+filename)

def readDataFromSheet(excelFile:ExcelFile, sheetName:str, startRow=0) -> DataFrame:
    """Reads data from the Excel Sheet
    
    Arguments:
        excelFile {ExcelFile} -- [description]
        sheetName {str} -- [description]
    
    Keyword Arguments:
        startRow {int} -- [description] (default: {0})
    
    Returns:
        DataFrame -- [description]
    """
    sheetData = excelFile.parse(sheetName, startRow)
    return sheetData

def consolidateSheetData(dateRecorded:datetime.date, sheetData:DataFrame) -> ExcelDataConsolidation:
    """Consolidates the Excel sheet data and creates an ExcelDataConsolidation object
    
    Returns:
        ExcelDataConsolidation -- consolidates sheet data
    """
    excelDataConsolidated = ExcelDataConsolidation()
    
    excelDataConsolidated.setDateRecorded(dateRecorded)
    excelDataConsolidated.setStartTime(sheetData[START_TIME_SHEET_IDENTIFIER][0])
    excelDataConsolidated.setLeavingTime(getLeavingTime(sheetData))
    
    workDuration = 0
    breakDuration = 0
    lunchDuration = 0
    for index in range(len(sheetData[DATE_RECORDED_SHEET_IDENTIFIER])):
        if (math.isnan(sheetData[DURATION_SHEET_IDENTIFIER][index])):
            continue
        if sheetData[CATEGORY_SHEET_IDENTIFIER][index].lower() == CATEGORY_CONST_DICT['work']:
            workDuration += sheetData[DURATION_SHEET_IDENTIFIER][index]/60
            projectDict = processProjectTaskDict(sheetData.loc[index], excelDataConsolidated.getProjectDict())
            excelDataConsolidated.setProjectDict(projectDict)
        else:
            breakDuration += (sheetData[DURATION_SHEET_IDENTIFIER][index])/60
            if sheetData[PROJECT_NAME_SHEET_IDENTIFIER][index].lower() == CATEGORY_CONST_DICT['lunch']:
                lunchDuration += (sheetData[DURATION_SHEET_IDENTIFIER][index])/60
    excelDataConsolidated.setWorkHours(workDuration)
    excelDataConsolidated.setBreakHours(breakDuration)
    excelDataConsolidated.setLunchDuration(lunchDuration)
    return excelDataConsolidated

def processProjectTaskDict(sheetData:DataFrame, projTaskDict:dict) -> dict:
    """Populates the specified task details into the current projectDict and returns a copy of the dictionary
    
    Arguments:
        sheetData {DataFrame} -- [description]
        projTaskDict {dict} -- [description]
    
    Returns:
        dict -- [description]
    """
    projectName = sheetData[PROJECT_NAME_SHEET_IDENTIFIER]
    taskName = sheetData[TASK_NAME_SHEET_IDENTIFIER]
    taskType = sheetData[TYPE_SHEET_IDENTIFIER]
    effort = sheetData[DURATION_SHEET_IDENTIFIER]
    description = sheetData[DESCRIPTION_SHEET_IDENTIFIER]
    projectDict = projTaskDict
    taskDict = dict()
    taskDict[DURATION_SHEET_IDENTIFIER] = 0
    taskDict[DESCRIPTION_SHEET_IDENTIFIER] = ""
    if projectName in projTaskDict:
        if taskName in projTaskDict[projectName]:
            taskDict = projTaskDict[projectName][taskName]
    else:
        projectDict[projectName] = {}
    taskDict[TYPE_SHEET_IDENTIFIER] = taskType
    taskDict[DURATION_SHEET_IDENTIFIER] += effort/60
    if (description.lower() not in taskDict[DESCRIPTION_SHEET_IDENTIFIER].lower()):
        taskDict[DESCRIPTION_SHEET_IDENTIFIER] += " "+description+"."
    projectDict[projectName][taskName] = taskDict
    return projectDict
    

def populateDailyActivityObject(excelDataConsolidated:ExcelDataConsolidation) -> list:
    """Creates and Populates a DailyActivity objects
    
    Arguments:
        excelDataConsolidated {ExcelDataConsolidation} -- [description]
    
    Returns:
        list -- [description]
    """
    dailyActivityList = []
    for projectName in excelDataConsolidated.getProjectDict():
        for taskName in excelDataConsolidated.getProjectDict()[projectName]:
            dailyActivity = DailyActivity()
            dailyActivity.setDate(excelDataConsolidated.getDateRecorded())
            dailyActivity.setTaskId(getTaskId(projectName, taskName))
            dailyActivity.setType(excelDataConsolidated.getProjectDict()[projectName][taskName][TYPE_SHEET_IDENTIFIER])
            dailyActivity.setEffort(excelDataConsolidated.getProjectDict()[projectName][taskName][DURATION_SHEET_IDENTIFIER])
            dailyActivity.setDescription(excelDataConsolidated.getProjectDict()[projectName][taskName][DESCRIPTION_SHEET_IDENTIFIER])
            dailyActivityList.append(dailyActivity)
    return dailyActivityList

def populateDayOverview(excelDataConsolidated:ExcelDataConsolidation, overviewSheetData) -> DayOverview:
    """Populates the Day Overview Object
    
    Arguments:
        excelDataConsolidated {ExcelDataConsolidation} -- [description]
        overviewSheetData {[type]} -- [description]
    
    Returns:
        DayOverview -- [description]
    """
    if (math.isnan(overviewSheetData["Prayer"][0])):
        prayer = PRAYER_DEFAULT
    else:
        prayer = not PRAYER_DEFAULT
    return DayOverview(excelDataConsolidated.getDateRecorded(), excelDataConsolidated.getStartTime(),
        excelDataConsolidated.getLeavingTime(), excelDataConsolidated.getWorkHours(), excelDataConsolidated.getBreakHours(),
        excelDataConsolidated.getLunchDuration(), getMorningScheduleId(overviewSheetData["Morning"][0]), prayer)

def getTaskId(projectName: str, taskName: str) -> int:
    """Get Task ID
    
    Arguments:
        projectName {str} -- [description]
        taskName {str} -- [description]
    
    Returns:
        int -- [description]
    """
    projectId = excelDatabase.getProjectId(projectName)
    taskId = excelDatabase.getTaskId(projectId, taskName)
    return taskId

def getMorningScheduleId(morningSchedule:str) -> int:
    """Get morningScheduleId corresponding the morningSchedule Indicator
    
    Arguments:
        morningSchedule {str} -- [description]
    
    Returns:
        int -- [description]
    """
    morningScheduleMappingDict = {"B": 1, "B+SL": 2, "BC": 3, "BC+SL": 4, "SL": 5}
    return morningScheduleMappingDict[morningSchedule]

def getDateRecordedFromFileName(filename:str) -> datetime.date:
    """Parses the date recorded from the file name
    
    Arguments:
        filename {str} -- [description]
    
    Returns:
        datetime.date -- [description]
    """
    dateField = filename.split(".")[0]
    dateComponents = list(map(int, dateField.split("-")))
    return datetime.date(dateComponents[2], dateComponents[0], dateComponents[1])

def saveDailyActivityList(dailyActivityList: list) -> list:
    """Persists the daily activity objects list into the Database
    
    Arguments:
        dailyActivityList {list} -- [description]
    
    Returns:
        list -- [description]
    """
    dailyActivityRecordList = excelDatabase.saveBulkDailyActivity(dailyActivityList)
    return dailyActivityRecordList

def saveDayOverview(dayOverview: DayOverview):
    """Persists the Day Overview record into the database
    
    Arguments:
        dayOverview {DayOverview} -- [description]
    
    Returns:
        [type] -- [description]
    """
    return excelDatabase.saveDayOverview(dayOverview)

def getExcelFileNamesToBeProcessed() -> list:
    """Gets the list of daily activity files that match
    following regex pattern mm-dd-yyyy.xlsx
    
    Returns:
        list -- [description]
    """
    fileList = os.listdir(FILE_PATH)
    taskFileNameList = [fileName for fileName in fileList if (configurationValues['fileExtension'] in fileName)
    and (not fileName.startswith("~$") and (re.match(configurationValues["taskSheetTitleRegex"], fileName)))]
    taskFileNameList.sort(key = parseUnixTime, reverse=True)
    return taskFileNameList

def saveDailyActivityListAndDayOverview(dailyActivityList: list, dayOverview: DayOverview):
    """Saves the Daily Activity List and the Day Overview data
    
    Arguments:
        dailyActivityList {list} -- [description]
        dayOverview {DayOverview} -- [description]
    
    Raises:
        Exception: [description]
        Exception: [description]
    
    Returns:
        [type] -- [description]
    """
    isDailyActivityListSaved = saveDailyActivityList(dailyActivityList)
    isDayOverviewSaved = saveDayOverview(dayOverview)
    excelDatabase.commit()
    if (isDailyActivityListSaved and isDayOverviewSaved):
        return True
    if (isDailyActivityListSaved):
        raise Exception("Could not save DayOverview record")
    raise Exception("Could not save DailyActivityList")

def createTaskTrackerWorkbook():
    """Creates the Daily Activity sheet
    """
    taskTrackerFile = load_workbook(os.path.abspath(TASK_TRACKER_TEMPLATE_PATH))
    fileName = datetime.datetime.now().strftime("%m-%d-%Y"+".xlsx")
    
    consolidationSheet = taskTrackerFile["Consolidated"]
    logSheet = taskTrackerFile["Sheet1"]
    projectTaskMapping = taskTrackerFile["Project-Task Mapping"]
    
    dataValidationDict = {}
    dataValidationDict[CELL_IDENTIFIER["taskName"]] = Constants.TASK_DATA_VALIDATION
    taskTypeList = excelDatabase.getTaskTypes()
    dataValidationDict[CELL_IDENTIFIER["taskType"]] = '"'+", ".join(taskTypeList)+'"'
    addDataValidation(logSheet, dataValidationDict)
    saveActiveProjectTaskMapping(projectTaskMapping)
    consolidationSheet.cell(row = 2, column = 6).value = datetime.datetime.now().strftime("%I:%M %p")
    taskTrackerFile.save(FILE_PATH+"/"+fileName)

def addDataValidation(worksheet:Worksheet, cellFormulaDict:dict):
    """Adds data validation to the specified worksheet
    on the listed cells mapped to the formula.
    
    Arguments:
        worksheet {Worksheet} -- [description]
        cellFormulaDict {dict} -- Cell address is the key and the formula is the value
    """
    if cellFormulaDict[CELL_IDENTIFIER["taskName"]] is not None:
        taskDataValidation = DataValidation(type="list", formula1=cellFormulaDict[CELL_IDENTIFIER["taskName"]], allow_blank=False)
        worksheet.add_data_validation(taskDataValidation)
        taskDataValidation.add(CELL_IDENTIFIER["taskName"])
    if cellFormulaDict[CELL_IDENTIFIER["taskType"]] is not None:
        taskTypeDataValidation = DataValidation(type="list", formula1=cellFormulaDict[CELL_IDENTIFIER["taskType"]], allow_blank=False)
        worksheet.add_data_validation(taskTypeDataValidation)
        taskTypeDataValidation.add(CELL_IDENTIFIER["taskType"])

def getDefaultProjects() -> list:
    """Creates default projects as defined in the configuration
    with default task "-".
    
    Returns:
        list -- List of Projects
    """
    projectList = []
    for projectName in DEFAULT_PROJECT:
        project = Project()
        project.projectName = projectName
        taskList = []
        task = Task()
        task.taskName = "-"
        taskList.append(task)
        project.taskList = taskList
        projectList.append(project)
    return projectList

def saveActiveProjectTaskMapping(worksheet:Worksheet, startRow:int = 1, startColumn:int = 1):
    """Saves the active project list and active task list data.
    The projects are stored in one row, the tasks corresponding to the project are saved in the same column as the project
    
    Arguments:
        worksheet {Worksheet} -- [description]
        startRow {int} -- starting row form where the data should be stored
        startColumn {int} -- starting column form where the data should be stored
    """
    projectList = excelDatabase.getActiveProjects()
    if projectList is None:
        return
    for project in projectList:
        taskList = excelDatabase.getActiveTasksByProjectId(project.projectId)
        if taskList is None:
            taskList = []
            task = Task()
            task.taskName = "-"
            taskList.append(task)
        project.taskList = taskList
    defaultProject = getDefaultProjects()
    projectList.extend(defaultProject)
    column = startColumn
    for project in projectList:
        row = startRow
        worksheet.cell(row=row, column=column).value = project.projectName
        for task in project.taskList:
            row += 1
            worksheet.cell(row=row, column=column).value = task.taskName
        column += 1

def getLeavingTime(sheetData: DataFrame) -> datetime:
    """Get the leaving time in the defined in the dataframe
    
    Arguments:
        sheetData {DataFrame} -- [description]
    
    Returns:
        datetime -- [description]
    """
    recordCount = len(sheetData[END_TIME_SHEET_IDENTIFIER])
    lastDuration = sheetData[DURATION_SHEET_IDENTIFIER][recordCount-1]
    if not math.isnan(lastDuration):
        return sheetData[END_TIME_SHEET_IDENTIFIER][recordCount-1]
    for index in range(recordCount-1, 0, -1):
        if math.isnan(sheetData[DURATION_SHEET_IDENTIFIER][index]):
            continue
        return sheetData[END_TIME_SHEET_IDENTIFIER][index]